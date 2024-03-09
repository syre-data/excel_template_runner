from typing import Optional
import os

import pandas
import openpyxl as xl
from openpyxl import utils as xl_utils
from openpyxl.utils import exceptions as xl_exceptions
from openpyxl.formula.tokenizer import Tokenizer, Token
from openpyxl.formula.translate import Translator
import formulas
import syre

from . import utils
from .types import (
    DataFormatType,
    DataFormatArgs,
    WorksheetId,
    ColumnId,
    Worksheet,
    ColumnSelection,
    HeaderAction,
    AssetFilter,
    AssetProperties,
    ReplaceRange,
)


def insert_data_from_excel(
    asset: syre.Asset,
    worksheet: Worksheet,
    data_worksheet: WorksheetId,
    column_selection: ColumnSelection,
    header_action: HeaderAction,
    current_column: int,
    skip_rows: int = 0,
    asset_path: Optional[str] = None,
):
    """Insert data into a worksheet from an Excel workbook.

    Args:
        asset (syre.Asset): Asset representing the data resource.
        worksheet (Worksheet): Template worksheet in which to insert the data.
        data_worksheet (WorksheetId): Worksheet id containing the data.
        column_selection (ColumnSelection): List of column headers or labels identifying the data to be copied into the template. [data]
        header_action (HeaderAction): How to label data. [template]
        current_column (int): Current column index of the template manipulation.
        skip_rows (int, optional): Number of rows to skip until the first header or data. Defaults to 0.
        asset_path (str): Relative path to the asset file.
            Used to label data if `data-label` is `HeaderAction.INSERT` or `HeaderAction.REPLACE`.
            Defaults to None.
    """
    asset_file_name = os.path.basename(asset.file)
    xl_model = formulas.ExcelModel().loads(asset.file).finish()
    calculated_data = xl_model.calculate()
    input_wb = xl.load_workbook(asset.file)

    if isinstance(data_worksheet, str):
        input_ws = input_wb.get_sheet_by_name(data_worksheet)
    elif isinstance(data_worksheet, int):
        input_ws = input_wb.worksheets[data_worksheet]
    else:
        raise TypeError("Invalid `data_worksheet`")

    current_column_excel = utils.index_to_excel(current_column)
    context_key = f"'[{asset_file_name}]{input_ws.title.upper()}'"
    worksheet.insert_cols(current_column_excel, len(column_selection))

    selection_type = utils.selection_type(column_selection)
    if selection_type is ColumnId.INDEX:
        input_data = [
            col
            for idx, col in enumerate(input_ws.iter_cols())
            if idx in column_selection
        ]
    elif selection_type is ColumnId.HEADER:
        raise NotImplementedError()
    else:
        raise TypeError("Invalid column selection")

    for input_column in input_data:
        data_row_start = 1
        if (
            header_action is HeaderAction.INSERT
            or header_action is HeaderAction.REPLACE
        ):
            worksheet.cell(row=1, column=current_column_excel).value = asset_path
            data_row_start += 1

        if header_action is HeaderAction.REPLACE:
            input_column = input_column[skip_rows:]

        for row, input_cell in enumerate(input_column, start=data_row_start):
            template_cell = worksheet.cell(row=row, column=current_column_excel)
            if type(input_cell.value) is str:
                tok = Tokenizer(input_cell.value)
                tokens = tok.items
                if len(tokens) == 0:
                    continue
                elif tokens[0].type == Token.LITERAL:
                    template_cell.value = input_cell.value
                else:
                    calculated_cell = calculated_data.get(
                        f"{context_key}!{input_cell.coordinate}"
                    )
                    template_cell.value = calculated_cell.value[0, 0]
            else:
                template_cell.value = input_cell.value


def insert_data_from_csv(
    asset: syre.Asset,
    worksheet: Worksheet,
    columns_selection: ColumnSelection,
    header_action: HeaderAction,
    current_column: int,
    skip_rows: int = 0,
    comment: Optional[str] = None,
    asset_path: Optional[str] = None,
):
    """Insert data into a worksheet from an Excel workbook.

    Args:
        asset (syre.Asset): Asset representing the data resource.
        worksheet (Worksheet): Template worksheet in which to insert the data.
        column_selection (ColumnSelection): List of column headers or labels identifying the data to be copied into the template. [data]
        header_action (HeaderAction): How to construct data headers. [template]
        current_column (int): Current column index of the template manipulation.
        skip_rows (int, optional): Number of rows to skip until the first header or data. Defaults to 0.
        comment (Optional[str (length 1)], optional): Comment character to ignore lines. Defaults to None.
        asset_path (str): Relative path to the asset file.
            Used to label data if `data-label` is `HeaderAction.INSERT` or `HeaderAction.REPLACE`.
            Defaults to None.
    """
    data = pandas.read_csv(asset.file, skiprows=skip_rows, comment=comment)
    data_selector_type = utils.selection_type(columns_selection)
    if data_selector_type is ColumnId.INDEX:
        df = data.iloc[:, columns_selection]
    elif data_selector_type is ColumnId.HEADER:
        raise NotImplementedError("todo")
        # df = data.loc[:, data_selector]
    else:
        raise RuntimeError("Invalid data selector")

    current_column_excel = utils.index_to_excel(current_column)
    worksheet.insert_cols(current_column_excel, len(columns_selection))
    for col_label, col in df.items():
        data_row_start = 1
        if (
            header_action is HeaderAction.INSERT
            or header_action is HeaderAction.REPLACE
        ):
            worksheet.cell(row=data_row_start, column=current_column_excel).value = (
                asset_path
            )
            data_row_start += 1

        if header_action is not HeaderAction.REPLACE:
            if type(col_label) is tuple:
                for row, label in enumerate(col_label, start=data_row_start):
                    worksheet.cell(row=row, column=current_column_excel).value = label

                data_row_start += len(col_label)
            else:
                worksheet.cell(
                    row=data_row_start, column=current_column_excel
                ).value = col_label
                data_row_start += 1

        for row, value in enumerate(col, start=data_row_start):
            worksheet.cell(row=row, column=current_column_excel, value=value)


def translate_tokenizer_formula(
    tokenizer: Tokenizer,
    replace_range: ReplaceRange,
    column_shift: int,
    header_action: HeaderAction,
    insertion_break_column: int,
):
    """Translates a Tokenizer's formula, adusting for column shifts.

    Args:
        tokenizer (Tokenizer):
        replace_range (int, int): Replaced columns in the template.
        column_shift (int): How many columns data was shifted due to insertions.
        header_action (HeaderAction): How data headers were inserted into the template.
        insertion_break_column (int): Column after the final data insertion column.
            e.g. If the last data was inserted in column B (index 2), this would be column C (index 3).

    Raises:
        ValueError: If formula range is invalid.
    """
    # NB: `column_shift` could be calculated from `replace-range` and `insertion_break_column`,
    # but is passed in for efficiency as this funciton may be called often and the column shift will be constant.
    for token in tokenizer.items:
        if token.type == Token.OPERAND and token.subtype == Token.RANGE:
            col_start, _, col_end, _ = xl_utils.range_boundaries(token.value)
            col_start = utils.excel_to_index(col_start)
            col_end = utils.excel_to_index(col_end)
            if col_end == replace_range[1]:
                # expand range to include new data
                formula_range = token.value.split(":")
                if len(formula_range) == 1:
                    token.value = Translator.translate_range(
                        formula_range[0], redelta=0, cdelta=column_shift
                    )
                elif len(formula_range) == 2:
                    formula_range[1] = Translator.translate_range(
                        formula_range[1], rdelta=0, cdelta=column_shift
                    )
                    token.value = ":".join(formula_range)
                else:
                    raise ValueError(f"Invalid range `{token.value}`")

            if col_start > replace_range[1]:
                token.value = Translator.translate_range(
                    token.value, rdelta=0, cdelta=column_shift
                )

            if header_action is HeaderAction.INSERT:
                formula_range = token.value.split(":")
                if len(formula_range) == 1:
                    if replace_range[0] <= col_start <= insertion_break_column:
                        try:
                            token.value = Translator.translate_range(
                                formula_range[0], rdelta=1, cdelta=0
                            )
                        except ValueError:
                            pass
                elif len(formula_range) == 2:
                    if (
                        replace_range[0] <= col_start <= insertion_break_column
                        and replace_range[0] <= col_end <= insertion_break_column
                    ):
                        try:
                            token.value = Translator.translate_range(
                                token.value, rdelta=1, cdelta=0
                            )
                        except ValueError:
                            pass
                else:
                    raise ValueError(f"Invalid range `{token.value}`")


def translate_worksheet_formulas(
    workbook: xl.Workbook,
    replace_range: ReplaceRange,
    insertion_break_column: int,
    header_action: HeaderAction,
):
    """Translates formulas that reference cells within the replaced ranged.

    Args:
        workbook (xl.Workbook): Template.
        replace_range (ReplaceRange): Replaced columns in the template.
        insertion_break_column (int): Column after the final data insertion column.
            e.g. If the last data was inserted in column B (index 2), this would be column C (index 3).
        column_shift (int): How many columns data was shifted due to insertions.
        header_action (HeaderAction): How data headers were inserted into the template.
    """
    column_shift = utils.column_shift(replace_range[1], insertion_break_column)
    if column_shift == 0 and header_action is not HeaderAction.INSERT:
        return

    for ws in workbook.worksheets:
        for col, column in enumerate(ws.iter_cols(), start=1):
            if replace_range[0] <= col <= insertion_break_column:
                # in replaced range or index
                continue

            for cell in column:
                if type(cell.value) is not str:
                    continue

                tok = Tokenizer(cell.value)
                tokens = tok.items
                if len(tokens) == 0 or tokens[0].type == Token.LITERAL:
                    continue

                translate_tokenizer_formula(
                    tok,
                    replace_range,
                    column_shift,
                    header_action,
                    insertion_break_column,
                )
                cell.value = tok.render()


def main(
    template_path: str,
    worksheet: WorksheetId,
    replace_range: (int, int),
    data_format_type: DataFormatType,
    column_selection: ColumnSelection,
    header_action: HeaderAction,
    output_path: str,
    data_format_args: DataFormatArgs = {},
    asset_filter: AssetFilter = {},
    output_properties: AssetProperties = {},
):
    """Use an Excel file as a template.

    Args:
        template_path (str): Aboslute path to the Excel template.
        worksheet (WorksheetId): Name or index of the worksheet. [template]
        replace_range (int, int): Column index range (0-based) to replace in the template. [template]
        data_format_type: (DataFormatType): Type of data format to expect.
        column_selection (ColumnSelection): List of column indexes (0-based) or labels identifying the data to be copied into the template. [data]
        header_action (HeaderAction): How to modify data headers when inserting.
        output_path (str): Relative path to the saved output file.
        data_format_args (DataFormatArgs): Arguments relevant for the expected data format.
        asset_filter (AssetFilter): Filter to search for input data.
        output_properties (AssetProperties): Asset properties assigned to the output Asset. Defaults to {}.

    Raises:
        ValueError: If `data_selector` is empty.
        RuntimeError: If `asset_filter` returns no assets.
    """
    if len(column_selection) == 0:
        raise ValueError("Empty data selector")

    db = syre.Database()
    template = xl.load_workbook(filename=template_path)
    ws = utils.get_worksheet(template, worksheet)

    delete_columns_count = replace_range[1] - replace_range[0] + 1
    ws.delete_cols(utils.index_to_excel(replace_range[0]), delete_columns_count)

    assets = db.find_assets(**asset_filter)
    if len(assets) == 0:
        raise RuntimeError("No matching assets")

    if header_action is HeaderAction.INSERT:
        ws.insert_rows(0)

    db_root_path = utils.canonicalize_db_root_path(db)
    current_col = replace_range[0]
    if data_format_type is DataFormatType.EXCEL_WORKBOOK:
        for asset in assets:
            asset_path = os.path.relpath(asset.file, db_root_path)
            insert_data_from_excel(
                asset,
                ws,
                data_format_args.sheet,
                column_selection,
                header_action,
                current_col,
                skip_rows=data_format_args.skip_rows,
                asset_path=asset_path,
            )
            current_col += 1

    elif data_format_type is DataFormatType.SPREADSHEET:
        for asset in assets:
            asset_path = os.path.relpath(asset.file, db_root_path)
            insert_data_from_csv(
                asset,
                ws,
                column_selection,
                header_action,
                current_col,
                skip_rows=data_format_args.skip_rows,
                comment=data_format_args.comment,
                asset_path=asset_path,
            )
            current_col += 1
    else:
        raise ValueError("Invalid data format type")

    translate_worksheet_formulas(
        template,
        replace_range,
        current_col,
        header_action,
    )

    path = db.add_asset(output_path, **output_properties)
    template.save(path)
